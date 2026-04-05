from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import bcrypt
import jwt as pyjwt
import secrets
import io
import json
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
import random

# PDF / Excel
from fpdf import FPDF
from openpyxl import Workbook

# ── Logging ──────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ── Database ─────────────────────────────────────────────
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ── App & Router ─────────────────────────────────────────
app = FastAPI(title="Bus Management System")
api = APIRouter(prefix="/api")

# ── JWT Config ───────────────────────────────────────────
JWT_ALGORITHM = "HS256"

def get_jwt_secret():
    return os.environ["JWT_SECRET"]

# ── Password helpers ─────────────────────────────────────
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

# ── Token helpers ────────────────────────────────────────
def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {"sub": user_id, "email": email, "role": role, "exp": datetime.now(timezone.utc) + timedelta(minutes=60), "type": "access"}
    return pyjwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return pyjwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

# ── Auth dependency ──────────────────────────────────────
async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = pyjwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ══════════════════════════════════════════════════════════
# PYDANTIC MODELS
# ══════════════════════════════════════════════════════════

class LoginReq(BaseModel):
    email: str
    password: str

class RegisterReq(BaseModel):
    email: str
    password: str
    name: str
    role: str = "vendor"

class ForgotPasswordReq(BaseModel):
    email: str

class ResetPasswordReq(BaseModel):
    token: str
    new_password: str

class TenderReq(BaseModel):
    tender_id: str
    pk_rate: float
    energy_rate: float
    subsidy_rate: float = 0
    subsidy_type: str = "per_km"
    description: str = ""
    status: str = "active"

class BusReq(BaseModel):
    bus_id: str
    bus_type: str = "12m_ac"
    capacity: int = 40
    tender_id: str = ""
    depot: str = ""
    status: str = "active"

class DriverReq(BaseModel):
    name: str
    license_number: str
    phone: str = ""
    bus_id: str = ""
    status: str = "active"

class EnergyReq(BaseModel):
    bus_id: str
    date: str
    units_charged: float
    tariff_rate: float = 10.0

class IncidentReq(BaseModel):
    incident_type: str
    description: str
    bus_id: str = ""
    driver_id: str = ""
    severity: str = "medium"

class DeductionRuleReq(BaseModel):
    name: str
    rule_type: str
    penalty_percent: float
    is_capped: bool = False
    cap_limit: float = 0
    description: str = ""
    active: bool = True

class SettingsReq(BaseModel):
    key: str
    value: str

class BillingGenerateReq(BaseModel):
    period_start: str
    period_end: str
    depot: str = ""

class TripDetail(BaseModel):
    trip_number: int
    start_time: str
    end_time: str
    direction: str = "outward"

class DutyReq(BaseModel):
    driver_license: str
    driver_name: str = ""
    driver_phone: str = ""
    bus_id: str
    route_name: str
    start_point: str
    end_point: str
    date: str
    trips: list = []

class InfractionReq(BaseModel):
    code: str
    category: str  # A-G
    description: str
    amount: float
    safety_flag: bool = False
    repeat_escalation: bool = True
    active: bool = True

class BillingWorkflowReq(BaseModel):
    invoice_id: str
    action: str  # submit, process, propose, depot_approve, regional_approve, rm_sanction, voucher, hq_approve, pay
    remarks: str = ""

class BusinessRuleReq(BaseModel):
    rule_key: str
    rule_value: str
    category: str = "general"
    description: str = ""

# ══════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════

@api.post("/auth/login")
async def login(req: LoginReq, request: Request, response: Response):
    email = req.email.lower().strip()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        lockout = attempt.get("last_attempt", datetime.now(timezone.utc)) + timedelta(minutes=15)
        if datetime.now(timezone.utc) < lockout:
            raise HTTPException(status_code=429, detail="Account locked. Try again in 15 minutes.")
        else:
            await db.login_attempts.delete_one({"identifier": identifier})
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(req.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"last_attempt": datetime.now(timezone.utc)}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")
    await db.login_attempts.delete_one({"identifier": identifier})
    uid = str(user["_id"])
    access = create_access_token(uid, email, user.get("role", "vendor"))
    refresh = create_refresh_token(uid)
    response.set_cookie(key="access_token", value=access, httponly=True, secure=False, samesite="lax", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return {"id": uid, "email": user["email"], "name": user.get("name", ""), "role": user.get("role", "vendor"), "token": access}

@api.post("/auth/register")
async def register(req: RegisterReq, response: Response):
    email = req.email.lower().strip()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    hashed = hash_password(req.password)
    doc = {"email": email, "password_hash": hashed, "name": req.name, "role": req.role, "created_at": datetime.now(timezone.utc).isoformat()}
    result = await db.users.insert_one(doc)
    uid = str(result.inserted_id)
    access = create_access_token(uid, email, req.role)
    refresh = create_refresh_token(uid)
    response.set_cookie(key="access_token", value=access, httponly=True, secure=False, samesite="lax", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return {"id": uid, "email": email, "name": req.name, "role": req.role, "token": access}

@api.get("/auth/me")
async def auth_me(user: dict = Depends(get_current_user)):
    return user

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@api.post("/auth/forgot-password")
async def forgot_password(req: ForgotPasswordReq):
    email = req.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user:
        return {"message": "If account exists, a reset link has been sent."}
    token = secrets.token_urlsafe(32)
    await db.password_reset_tokens.insert_one({
        "token": token, "email": email,
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
        "used": False
    })
    logger.info(f"Password reset token for {email}: {token}")
    return {"message": "If account exists, a reset link has been sent.", "reset_token": token}

@api.post("/auth/reset-password")
async def reset_password(req: ResetPasswordReq):
    record = await db.password_reset_tokens.find_one({"token": req.token, "used": False})
    if not record or record["expires_at"].replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Invalid or expired token")
    hashed = hash_password(req.new_password)
    await db.users.update_one({"email": record["email"]}, {"$set": {"password_hash": hashed}})
    await db.password_reset_tokens.update_one({"token": req.token}, {"$set": {"used": True}})
    return {"message": "Password reset successfully"}

@api.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = pyjwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        uid = str(user["_id"])
        access = create_access_token(uid, user["email"], user.get("role", "vendor"))
        response.set_cookie(key="access_token", value=access, httponly=True, secure=False, samesite="lax", max_age=3600, path="/")
        return {"message": "Token refreshed"}
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ══════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════

@api.get("/dashboard")
async def get_dashboard(date_from: str = "", date_to: str = "", depot: str = "", user: dict = Depends(get_current_user)):
    query = {}
    if depot:
        query["depot"] = depot
    # Buses — only needed fields
    buses = await db.buses.find(query, {"_id": 0, "bus_id": 1, "status": 1, "depot": 1}).to_list(1000)
    total_buses = len(buses)
    active_buses = len([b for b in buses if b.get("status") == "active"])
    # Drivers — count only
    total_drivers = await db.drivers.count_documents({})
    active_drivers = await db.drivers.count_documents({"status": "active"})
    # Trips — aggregation for totals + chart
    trip_match = {}
    if date_from and date_to:
        trip_match["date"] = {"$gte": date_from, "$lte": date_to}
    elif date_from:
        trip_match["date"] = {"$gte": date_from}
    trip_agg = await db.trip_data.aggregate([
        {"$match": trip_match},
        {"$group": {"_id": "$date", "actual_km": {"$sum": "$actual_km"}, "scheduled_km": {"$sum": "$scheduled_km"}}}
    ]).to_list(500)
    total_km = sum(d["actual_km"] for d in trip_agg)
    scheduled_km = sum(d["scheduled_km"] for d in trip_agg)
    km_chart = sorted([{"date": d["_id"], "actual_km": d["actual_km"], "scheduled_km": d["scheduled_km"]} for d in trip_agg], key=lambda x: x["date"])[-30:]
    # Energy — aggregation
    energy_match = {}
    if date_from and date_to:
        energy_match["date"] = {"$gte": date_from, "$lte": date_to}
    energy_agg = await db.energy_data.aggregate([
        {"$match": energy_match},
        {"$group": {"_id": "$date", "units": {"$sum": "$units_charged"}}}
    ]).to_list(500)
    total_energy = sum(d["units"] for d in energy_agg)
    energy_chart = sorted([{"date": d["_id"], "units": d["units"]} for d in energy_agg], key=lambda x: x["date"])[-30:]
    # Incidents — count only
    active_incidents = await db.incidents.count_documents({"status": {"$ne": "resolved"}})
    # Billing — aggregation
    billing_agg = await db.billing.aggregate([
        {"$group": {"_id": None, "total": {"$sum": "$final_payable"}}}
    ]).to_list(1)
    total_revenue = billing_agg[0]["total"] if billing_agg else 0
    # Revenue — aggregation
    rev_match = {}
    if date_from and date_to:
        rev_match["date"] = {"$gte": date_from, "$lte": date_to}
    rev_agg = await db.revenue_data.aggregate([
        {"$match": rev_match},
        {"$group": {"_id": None, "revenue": {"$sum": "$revenue_amount"}, "passengers": {"$sum": "$passengers"}}}
    ]).to_list(1)
    total_ticket_revenue = rev_agg[0]["revenue"] if rev_agg else 0
    total_passengers = rev_agg[0]["passengers"] if rev_agg else 0
    # Depot list
    depots = list(set(b.get("depot", "") for b in buses if b.get("depot")))
    return {
        "total_buses": total_buses, "active_buses": active_buses,
        "total_drivers": total_drivers, "active_drivers": active_drivers,
        "total_km": round(total_km, 2), "scheduled_km": round(scheduled_km, 2),
        "total_energy": round(total_energy, 2), "active_incidents": active_incidents,
        "total_revenue": round(total_revenue, 2),
        "total_ticket_revenue": round(total_ticket_revenue, 2),
        "total_passengers": total_passengers,
        "availability_pct": round((total_km / scheduled_km * 100) if scheduled_km > 0 else 0, 1),
        "km_chart": km_chart, "energy_chart": energy_chart, "depots": depots
    }

# ══════════════════════════════════════════════════════════
# TENDERS
# ══════════════════════════════════════════════════════════

@api.get("/tenders")
async def list_tenders(user: dict = Depends(get_current_user)):
    tenders = await db.tenders.find({}, {"_id": 0}).to_list(1000)
    return tenders

@api.post("/tenders")
async def create_tender(req: TenderReq, user: dict = Depends(get_current_user)):
    existing = await db.tenders.find_one({"tender_id": req.tender_id})
    if existing:
        raise HTTPException(status_code=400, detail="Tender ID already exists")
    doc = req.model_dump()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.tenders.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/tenders/{tender_id}")
async def update_tender(tender_id: str, req: TenderReq, user: dict = Depends(get_current_user)):
    update = req.model_dump()
    update.pop("tender_id", None)
    result = await db.tenders.update_one({"tender_id": tender_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tender not found")
    return {"message": "Tender updated"}

@api.delete("/tenders/{tender_id}")
async def delete_tender(tender_id: str, user: dict = Depends(get_current_user)):
    buses = await db.buses.find_one({"tender_id": tender_id})
    if buses:
        raise HTTPException(status_code=400, detail="Cannot delete: buses are assigned to this tender")
    result = await db.tenders.delete_one({"tender_id": tender_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tender not found")
    return {"message": "Tender deleted"}

# ══════════════════════════════════════════════════════════
# BUSES
# ══════════════════════════════════════════════════════════

@api.get("/buses")
async def list_buses(user: dict = Depends(get_current_user)):
    buses = await db.buses.find({}, {"_id": 0}).to_list(1000)
    return buses

@api.post("/buses")
async def create_bus(req: BusReq, user: dict = Depends(get_current_user)):
    existing = await db.buses.find_one({"bus_id": req.bus_id})
    if existing:
        raise HTTPException(status_code=400, detail="Bus ID already exists")
    doc = req.model_dump()
    kwh_map = {"12m_ac": 1.3, "9m_ac": 1.0, "12m_non_ac": 1.1, "9m_non_ac": 0.8}
    doc["kwh_per_km"] = kwh_map.get(req.bus_type, 1.0)
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.buses.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/buses/{bus_id}")
async def update_bus(bus_id: str, req: BusReq, user: dict = Depends(get_current_user)):
    update = req.model_dump()
    update.pop("bus_id", None)
    kwh_map = {"12m_ac": 1.3, "9m_ac": 1.0, "12m_non_ac": 1.1, "9m_non_ac": 0.8}
    update["kwh_per_km"] = kwh_map.get(req.bus_type, 1.0)
    result = await db.buses.update_one({"bus_id": bus_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bus not found")
    return {"message": "Bus updated"}

@api.delete("/buses/{bus_id}")
async def delete_bus(bus_id: str, user: dict = Depends(get_current_user)):
    result = await db.buses.delete_one({"bus_id": bus_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bus not found")
    return {"message": "Bus deleted"}

@api.put("/buses/{bus_id}/assign-tender")
async def assign_tender_to_bus(bus_id: str, tender_id: str = Query(...), user: dict = Depends(get_current_user)):
    tender = await db.tenders.find_one({"tender_id": tender_id})
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    result = await db.buses.update_one({"bus_id": bus_id}, {"$set": {"tender_id": tender_id}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bus not found")
    return {"message": "Tender assigned to bus"}

@api.get("/buses/{bus_id}")
async def get_bus(bus_id: str, user: dict = Depends(get_current_user)):
    bus = await db.buses.find_one({"bus_id": bus_id}, {"_id": 0})
    if not bus:
        raise HTTPException(status_code=404, detail="Bus not found")
    trips = await db.trip_data.find({"bus_id": bus_id}, {"_id": 0}).to_list(100)
    energy = await db.energy_data.find({"bus_id": bus_id}, {"_id": 0}).to_list(100)
    return {**bus, "trips": trips, "energy": energy}

# ══════════════════════════════════════════════════════════
# DRIVERS
# ══════════════════════════════════════════════════════════

@api.get("/drivers")
async def list_drivers(user: dict = Depends(get_current_user)):
    drivers = await db.drivers.find({}, {"_id": 0}).to_list(1000)
    return drivers

@api.post("/drivers")
async def create_driver(req: DriverReq, user: dict = Depends(get_current_user)):
    existing = await db.drivers.find_one({"license_number": req.license_number})
    if existing:
        raise HTTPException(status_code=400, detail="License number already exists")
    doc = req.model_dump()
    doc["id"] = str(uuid.uuid4())[:8]
    doc["performance_score"] = 100.0
    doc["penalties"] = []
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.drivers.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/drivers/{license_number}")
async def update_driver(license_number: str, req: DriverReq, user: dict = Depends(get_current_user)):
    update = req.model_dump()
    update.pop("license_number", None)
    result = await db.drivers.update_one({"license_number": license_number}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Driver not found")
    return {"message": "Driver updated"}

@api.delete("/drivers/{license_number}")
async def delete_driver(license_number: str, user: dict = Depends(get_current_user)):
    result = await db.drivers.delete_one({"license_number": license_number})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Driver not found")
    return {"message": "Driver deleted"}

@api.put("/drivers/{license_number}/assign-bus")
async def assign_bus_to_driver(license_number: str, bus_id: str = Query(...), user: dict = Depends(get_current_user)):
    bus = await db.buses.find_one({"bus_id": bus_id})
    if not bus:
        raise HTTPException(status_code=404, detail="Bus not found")
    result = await db.drivers.update_one({"license_number": license_number}, {"$set": {"bus_id": bus_id}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Driver not found")
    return {"message": "Bus assigned to driver"}

@api.get("/drivers/{license_number}/performance")
async def get_driver_performance(license_number: str, user: dict = Depends(get_current_user)):
    driver = await db.drivers.find_one({"license_number": license_number}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Driver not found")
    trips = await db.trip_data.find({"driver_id": license_number}, {"_id": 0}).to_list(1000)
    total_km = sum(t.get("actual_km", 0) for t in trips)
    total_trips = len(trips)
    incidents = await db.incidents.find({"driver_id": license_number}, {"_id": 0}).to_list(100)
    return {
        "driver": driver, "total_km": round(total_km, 2), "total_trips": total_trips,
        "incidents": len(incidents), "performance_score": driver.get("performance_score", 100)
    }

# ══════════════════════════════════════════════════════════
# LIVE OPERATIONS
# ══════════════════════════════════════════════════════════

@api.get("/live-operations")
async def get_live_operations(user: dict = Depends(get_current_user)):
    buses = await db.buses.find({"status": "active"}, {"_id": 0}).to_list(1000)
    live_data = []
    center_lat, center_lng = 17.385, 78.486
    for bus in buses:
        lat = center_lat + random.uniform(-0.05, 0.05)
        lng = center_lng + random.uniform(-0.05, 0.05)
        speed = random.randint(15, 60)
        status = random.choice(["on_route", "on_route", "on_route", "at_stop", "charging"])
        live_data.append({
            "bus_id": bus["bus_id"], "bus_type": bus.get("bus_type", ""),
            "lat": round(lat, 6), "lng": round(lng, 6),
            "speed": speed, "status": status,
            "driver": bus.get("driver_name", ""),
            "depot": bus.get("depot", "")
        })
    return live_data

@api.get("/live-operations/alerts")
async def get_alerts(user: dict = Depends(get_current_user)):
    alerts = []
    buses = await db.buses.find({"status": "active"}, {"_id": 0}).to_list(1000)
    alert_types = ["GPS Failure", "Overspeeding", "Route Deviation", "Late Departure", "Low Battery", "PIS Failure"]
    for _ in range(min(5, len(buses))):
        bus = random.choice(buses)
        alerts.append({
            "id": str(uuid.uuid4())[:8],
            "bus_id": bus["bus_id"],
            "alert_type": random.choice(alert_types),
            "severity": random.choice(["low", "medium", "high"]),
            "timestamp": (datetime.now(timezone.utc) - timedelta(minutes=random.randint(1, 120))).isoformat(),
            "resolved": random.choice([True, False, False])
        })
    return alerts

# ══════════════════════════════════════════════════════════
# ENERGY MANAGEMENT
# ══════════════════════════════════════════════════════════

@api.get("/energy")
async def list_energy(date_from: str = "", date_to: str = "", bus_id: str = "", user: dict = Depends(get_current_user)):
    query = {}
    if bus_id:
        query["bus_id"] = bus_id
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    data = await db.energy_data.find(query, {"_id": 0}).to_list(3000)
    return data

@api.post("/energy")
async def add_energy(req: EnergyReq, user: dict = Depends(get_current_user)):
    doc = req.model_dump()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.energy_data.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.get("/energy/report")
async def energy_report(date_from: str = "", date_to: str = "", user: dict = Depends(get_current_user)):
    query = {}
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    data = await db.energy_data.find(query, {"_id": 0}).to_list(3000)
    buses = await db.buses.find({}, {"_id": 0}).to_list(1000)
    bus_map = {b["bus_id"]: b for b in buses}
    trips = await db.trip_data.find(query, {"_id": 0}).to_list(3000)
    bus_km = {}
    for t in trips:
        bid = t.get("bus_id", "")
        bus_km[bid] = bus_km.get(bid, 0) + t.get("actual_km", 0)
    bus_energy = {}
    for e in data:
        bid = e.get("bus_id", "")
        if bid not in bus_energy:
            bus_energy[bid] = {"bus_id": bid, "actual_kwh": 0, "tariff": e.get("tariff_rate", 10)}
        bus_energy[bid]["actual_kwh"] += e.get("units_charged", 0)
    report = []
    for bid, ed in bus_energy.items():
        bus = bus_map.get(bid, {})
        kwh_per_km = bus.get("kwh_per_km", 1.0)
        km = bus_km.get(bid, 0)
        allowed = km * kwh_per_km
        actual = ed["actual_kwh"]
        tariff = ed["tariff"]
        report.append({
            "bus_id": bid, "bus_type": bus.get("bus_type", ""),
            "km_operated": round(km, 2), "kwh_per_km": kwh_per_km,
            "allowed_kwh": round(allowed, 2), "actual_kwh": round(actual, 2),
            "efficiency": round((actual / allowed * 100) if allowed > 0 else 0, 1),
            "allowed_cost": round(allowed * tariff, 2),
            "actual_cost": round(actual * tariff, 2),
            "adjustment": round(min(actual, allowed) * tariff, 2)
        })
    total_allowed = sum(r["allowed_kwh"] for r in report)
    total_actual = sum(r["actual_kwh"] for r in report)
    return {
        "report": report,
        "summary": {
            "total_allowed_kwh": round(total_allowed, 2),
            "total_actual_kwh": round(total_actual, 2),
            "total_efficiency": round((total_actual / total_allowed * 100) if total_allowed > 0 else 0, 1)
        }
    }

# ══════════════════════════════════════════════════════════
# KPI
# ══════════════════════════════════════════════════════════

@api.get("/kpi")
async def get_kpi(date_from: str = "", date_to: str = "", user: dict = Depends(get_current_user)):
    query = {}
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    trips = await db.trip_data.find(query, {"_id": 0}).to_list(3000)
    energy = await db.energy_data.find(query, {"_id": 0}).to_list(3000)
    buses = await db.buses.find({}, {"_id": 0}).to_list(1000)
    incidents = await db.incidents.find({}, {"_id": 0}).to_list(1000)
    total_scheduled = sum(t.get("scheduled_km", 0) for t in trips)
    total_actual = sum(t.get("actual_km", 0) for t in trips)
    total_energy = sum(e.get("units_charged", 0) for e in energy)
    active_buses = len([b for b in buses if b.get("status") == "active"])
    return {
        "fleet_availability": round((total_actual / total_scheduled * 100) if total_scheduled > 0 else 0, 1),
        "km_efficiency": round((total_actual / total_scheduled * 100) if total_scheduled > 0 else 0, 1),
        "energy_per_km": round((total_energy / total_actual) if total_actual > 0 else 0, 3),
        "total_km_operated": round(total_actual, 2),
        "total_scheduled_km": round(total_scheduled, 2),
        "total_energy_consumed": round(total_energy, 2),
        "active_fleet": active_buses,
        "total_incidents": len(incidents),
        "open_incidents": len([i for i in incidents if i.get("status") != "resolved"]),
        "avg_speed": round(random.uniform(28, 35), 1),
        "on_time_pct": round(random.uniform(85, 95), 1)
    }

# ══════════════════════════════════════════════════════════
# DEDUCTION ENGINE
# ══════════════════════════════════════════════════════════

@api.get("/deductions/rules")
async def list_rules(user: dict = Depends(get_current_user)):
    rules = await db.deduction_rules.find({}, {"_id": 0}).to_list(100)
    return rules

@api.post("/deductions/rules")
async def create_rule(req: DeductionRuleReq, user: dict = Depends(get_current_user)):
    doc = req.model_dump()
    doc["id"] = str(uuid.uuid4())[:8]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.deduction_rules.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/deductions/rules/{rule_id}")
async def update_rule(rule_id: str, req: DeductionRuleReq, user: dict = Depends(get_current_user)):
    update = req.model_dump()
    result = await db.deduction_rules.update_one({"id": rule_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    return {"message": "Rule updated"}

@api.delete("/deductions/rules/{rule_id}")
async def delete_rule(rule_id: str, user: dict = Depends(get_current_user)):
    result = await db.deduction_rules.delete_one({"id": rule_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    return {"message": "Rule deleted"}

@api.post("/deductions/apply")
async def apply_deductions(period_start: str = Query(...), period_end: str = Query(...), user: dict = Depends(get_current_user)):
    rules = await db.deduction_rules.find({"active": True}, {"_id": 0}).to_list(100)
    trips = await db.trip_data.find({"date": {"$gte": period_start, "$lte": period_end}}, {"_id": 0}).to_list(3000)
    tenders = await db.tenders.find({}, {"_id": 0}).to_list(100)
    tender_map = {t["tender_id"]: t for t in tenders}
    buses = await db.buses.find({}, {"_id": 0}).to_list(1000)
    bus_map = {b["bus_id"]: b for b in buses}
    total_scheduled = sum(t.get("scheduled_km", 0) for t in trips)
    total_actual = sum(t.get("actual_km", 0) for t in trips)
    missed_km = max(0, total_scheduled - total_actual)
    avg_pk_rate = 0
    if tenders:
        avg_pk_rate = sum(t.get("pk_rate", 0) for t in tenders) / len(tenders)
    base_payment = total_actual * avg_pk_rate
    availability_deduction = missed_km * avg_pk_rate
    performance_deduction = 0
    system_deduction = 0
    capped_total = 0
    uncapped_total = 0
    breakdown = []
    for rule in rules:
        rt = rule.get("rule_type", "")
        pct = rule.get("penalty_percent", 0)
        amount = base_payment * (pct / 100)
        if rule.get("is_capped") and rule.get("cap_limit", 0) > 0:
            amount = min(amount, rule["cap_limit"])
            capped_total += amount
        else:
            uncapped_total += amount
        if rt == "performance":
            performance_deduction += amount
        elif rt == "system":
            system_deduction += amount
        breakdown.append({"rule": rule["name"], "type": rt, "percent": pct, "amount": round(amount, 2)})
    total_deduction = availability_deduction + performance_deduction + system_deduction
    return {
        "period": {"start": period_start, "end": period_end},
        "base_payment": round(base_payment, 2),
        "missed_km": round(missed_km, 2),
        "availability_deduction": round(availability_deduction, 2),
        "performance_deduction": round(performance_deduction, 2),
        "system_deduction": round(system_deduction, 2),
        "total_deduction": round(total_deduction, 2),
        "breakdown": breakdown
    }

# ══════════════════════════════════════════════════════════
# BILLING
# ══════════════════════════════════════════════════════════

@api.post("/billing/generate")
async def generate_invoice(req: BillingGenerateReq, user: dict = Depends(get_current_user)):
    period_start = req.period_start
    period_end = req.period_end
    depot = req.depot
    bus_query = {}
    if depot:
        bus_query["depot"] = depot
    buses = await db.buses.find(bus_query, {"_id": 0}).to_list(1000)
    bus_ids = [b["bus_id"] for b in buses]
    bus_map = {b["bus_id"]: b for b in buses}
    tenders = await db.tenders.find({}, {"_id": 0}).to_list(100)
    tender_map = {t["tender_id"]: t for t in tenders}
    trip_query = {"date": {"$gte": period_start, "$lte": period_end}}
    if bus_ids:
        trip_query["bus_id"] = {"$in": bus_ids}
    trips = await db.trip_data.find(trip_query, {"_id": 0}).to_list(3000)
    energy_query = {"date": {"$gte": period_start, "$lte": period_end}}
    if bus_ids:
        energy_query["bus_id"] = {"$in": bus_ids}
    energy = await db.energy_data.find(energy_query, {"_id": 0}).to_list(3000)
    # Step 1: Total KM
    total_km = sum(t.get("actual_km", 0) for t in trips)
    scheduled_km = sum(t.get("scheduled_km", 0) for t in trips)
    # Step 2: PK Rate (weighted average by bus tender)
    bus_km = {}
    for t in trips:
        bid = t.get("bus_id", "")
        bus_km[bid] = bus_km.get(bid, 0) + t.get("actual_km", 0)
    weighted_pk = 0
    for bid, km in bus_km.items():
        bus = bus_map.get(bid, {})
        tender = tender_map.get(bus.get("tender_id", ""), {})
        pk_rate = tender.get("pk_rate", 0)
        weighted_pk += km * pk_rate
    base_payment = weighted_pk
    avg_pk_rate = weighted_pk / total_km if total_km > 0 else 0
    # Step 3: Energy
    bus_energy = {}
    for e in energy:
        bid = e.get("bus_id", "")
        if bid not in bus_energy:
            bus_energy[bid] = {"actual": 0, "tariff": e.get("tariff_rate", 10)}
        bus_energy[bid]["actual"] += e.get("units_charged", 0)
    total_allowed_energy = 0
    total_actual_energy = 0
    tariff_rate = 10
    for bid in set(list(bus_km.keys()) + list(bus_energy.keys())):
        bus = bus_map.get(bid, {})
        kwh_per_km = bus.get("kwh_per_km", 1.0)
        km = bus_km.get(bid, 0)
        allowed = km * kwh_per_km
        actual = bus_energy.get(bid, {}).get("actual", 0)
        tariff_rate = bus_energy.get(bid, {}).get("tariff", 10)
        total_allowed_energy += allowed
        total_actual_energy += actual
    allowed_cost = total_allowed_energy * tariff_rate
    actual_cost = total_actual_energy * tariff_rate
    energy_adjustment = min(actual_cost, allowed_cost)
    # Step 4: Subsidy
    subsidy = 0
    for bid, km in bus_km.items():
        bus = bus_map.get(bid, {})
        tender = tender_map.get(bus.get("tender_id", ""), {})
        sr = tender.get("subsidy_rate", 0)
        st = tender.get("subsidy_type", "per_km")
        if st == "per_km":
            subsidy += km * sr
        elif st == "per_bus":
            subsidy += sr
    # Step 5: Deductions
    missed_km = max(0, scheduled_km - total_km)
    availability_deduction = missed_km * avg_pk_rate
    rules = await db.deduction_rules.find({"active": True}, {"_id": 0}).to_list(100)
    performance_deduction = 0
    system_deduction = 0
    for rule in rules:
        rt = rule.get("rule_type", "")
        pct = rule.get("penalty_percent", 0)
        amount = base_payment * (pct / 100)
        if rule.get("is_capped") and rule.get("cap_limit", 0) > 0:
            amount = min(amount, rule["cap_limit"])
        if rt == "performance":
            performance_deduction += amount
        elif rt == "system":
            system_deduction += amount
    total_deduction = availability_deduction + performance_deduction + system_deduction
    # Step 6: Final
    final_payable = base_payment + energy_adjustment + subsidy - total_deduction
    invoice = {
        "invoice_id": f"INV-{str(uuid.uuid4())[:8].upper()}",
        "period_start": period_start, "period_end": period_end,
        "depot": depot or "All",
        "total_km": round(total_km, 2), "scheduled_km": round(scheduled_km, 2),
        "avg_pk_rate": round(avg_pk_rate, 2), "base_payment": round(base_payment, 2),
        "allowed_energy_kwh": round(total_allowed_energy, 2),
        "actual_energy_kwh": round(total_actual_energy, 2),
        "tariff_rate": tariff_rate,
        "allowed_energy_cost": round(allowed_cost, 2),
        "actual_energy_cost": round(actual_cost, 2),
        "energy_adjustment": round(energy_adjustment, 2),
        "subsidy": round(subsidy, 2),
        "missed_km": round(missed_km, 2),
        "availability_deduction": round(availability_deduction, 2),
        "performance_deduction": round(performance_deduction, 2),
        "system_deduction": round(system_deduction, 2),
        "total_deduction": round(total_deduction, 2),
        "final_payable": round(final_payable, 2),
        "status": "draft",
        "workflow_state": "draft",
        "workflow_log": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.billing.insert_one(dict(invoice))
    invoice.pop("_id", None)
    return invoice

@api.get("/billing")
async def list_invoices(user: dict = Depends(get_current_user)):
    invoices = await db.billing.find({}, {"_id": 0}).to_list(1000)
    return invoices

@api.get("/billing/{invoice_id}")
async def get_invoice(invoice_id: str, user: dict = Depends(get_current_user)):
    inv = await db.billing.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return inv

@api.get("/billing/{invoice_id}/export-pdf")
async def export_invoice_pdf(invoice_id: str, user: dict = Depends(get_current_user)):
    inv = await db.billing.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(190, 10, "TGSRTC - Bus Management Invoice", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(190, 6, f"Invoice ID: {inv['invoice_id']}", ln=True)
    pdf.cell(190, 6, f"Period: {inv['period_start']} to {inv['period_end']}", ln=True)
    pdf.cell(190, 6, f"Depot: {inv.get('depot', 'All')}", ln=True)
    pdf.cell(190, 6, f"Generated: {inv.get('created_at', '')[:10]}", ln=True)
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(190, 8, "Billing Summary", ln=True)
    pdf.set_font("Helvetica", "", 10)
    rows = [
        ("Total KM Operated", f"{inv['total_km']:,.2f} km"),
        ("Scheduled KM", f"{inv['scheduled_km']:,.2f} km"),
        ("Avg PK Rate", f"Rs. {inv['avg_pk_rate']:,.2f}/km"),
        ("Base Payment (KM x PK Rate)", f"Rs. {inv['base_payment']:,.2f}"),
        ("", ""),
        ("Allowed Energy", f"{inv['allowed_energy_kwh']:,.2f} kWh"),
        ("Actual Energy", f"{inv['actual_energy_kwh']:,.2f} kWh"),
        ("Tariff Rate", f"Rs. {inv['tariff_rate']:,.2f}/kWh"),
        ("Energy Adjustment", f"Rs. {inv['energy_adjustment']:,.2f}"),
        ("", ""),
        ("Subsidy", f"Rs. {inv['subsidy']:,.2f}"),
        ("", ""),
        ("Missed KM", f"{inv['missed_km']:,.2f} km"),
        ("Availability Deduction", f"Rs. {inv['availability_deduction']:,.2f}"),
        ("Performance Deduction", f"Rs. {inv['performance_deduction']:,.2f}"),
        ("System Deduction", f"Rs. {inv['system_deduction']:,.2f}"),
        ("Total Deductions", f"Rs. {inv['total_deduction']:,.2f}"),
    ]
    for label, val in rows:
        if label == "":
            pdf.ln(2)
            continue
        pdf.cell(110, 6, label)
        pdf.cell(80, 6, val, ln=True, align="R")
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(110, 8, "FINAL PAYABLE")
    pdf.cell(80, 8, f"Rs. {inv['final_payable']:,.2f}", ln=True, align="R")
    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={inv['invoice_id']}.pdf"})

@api.get("/billing/{invoice_id}/export-excel")
async def export_invoice_excel(invoice_id: str, user: dict = Depends(get_current_user)):
    inv = await db.billing.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.append(["TGSRTC Bus Management Invoice"])
    ws.append([])
    ws.append(["Invoice ID", inv["invoice_id"]])
    ws.append(["Period", f"{inv['period_start']} to {inv['period_end']}"])
    ws.append(["Depot", inv.get("depot", "All")])
    ws.append([])
    ws.append(["Component", "Value"])
    fields = [
        ("Total KM", inv["total_km"]), ("Scheduled KM", inv["scheduled_km"]),
        ("Avg PK Rate (Rs/km)", inv["avg_pk_rate"]), ("Base Payment", inv["base_payment"]),
        ("Allowed Energy (kWh)", inv["allowed_energy_kwh"]), ("Actual Energy (kWh)", inv["actual_energy_kwh"]),
        ("Tariff (Rs/kWh)", inv["tariff_rate"]), ("Energy Adjustment", inv["energy_adjustment"]),
        ("Subsidy", inv["subsidy"]),
        ("Missed KM", inv["missed_km"]), ("Availability Deduction", inv["availability_deduction"]),
        ("Performance Deduction", inv["performance_deduction"]), ("System Deduction", inv["system_deduction"]),
        ("Total Deductions", inv["total_deduction"]),
        ("FINAL PAYABLE", inv["final_payable"])
    ]
    for label, val in fields:
        ws.append([label, val])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={inv['invoice_id']}.xlsx"})

# ══════════════════════════════════════════════════════════
# REPORTS
# ══════════════════════════════════════════════════════════

@api.get("/reports")
async def generate_report(report_type: str = "operations", date_from: str = "", date_to: str = "", user: dict = Depends(get_current_user)):
    query = {}
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    if report_type == "operations":
        trips = await db.trip_data.find(query, {"_id": 0}).to_list(3000)
        return {"type": "operations", "data": trips, "count": len(trips)}
    elif report_type == "energy":
        data = await db.energy_data.find(query, {"_id": 0}).to_list(3000)
        return {"type": "energy", "data": data, "count": len(data)}
    elif report_type == "incidents":
        data = await db.incidents.find({}, {"_id": 0}).to_list(1000)
        return {"type": "incidents", "data": data, "count": len(data)}
    elif report_type == "billing":
        data = await db.billing.find({}, {"_id": 0}).to_list(1000)
        return {"type": "billing", "data": data, "count": len(data)}
    return {"type": report_type, "data": [], "count": 0}

@api.get("/reports/download")
async def download_report(report_type: str = "operations", date_from: str = "", date_to: str = "", fmt: str = "excel", user: dict = Depends(get_current_user)):
    query = {}
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    if report_type == "operations":
        data = await db.trip_data.find(query, {"_id": 0}).to_list(3000)
        cols = ["bus_id", "driver_id", "date", "scheduled_km", "actual_km"]
    elif report_type == "energy":
        data = await db.energy_data.find(query, {"_id": 0}).to_list(3000)
        cols = ["bus_id", "date", "units_charged", "tariff_rate"]
    elif report_type == "incidents":
        data = await db.incidents.find({}, {"_id": 0}).to_list(1000)
        cols = ["id", "incident_type", "bus_id", "severity", "status", "created_at"]
    elif report_type == "billing":
        data = await db.billing.find({}, {"_id": 0}).to_list(1000)
        cols = ["invoice_id", "period_start", "period_end", "base_payment", "energy_adjustment", "total_deduction", "final_payable"]
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")
    if fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.capitalize()
        ws.append(cols)
        for row in data:
            ws.append([row.get(c, "") for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={report_type}_report.xlsx"})
    else:
        pdf = FPDF()
        pdf.add_page("L")
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(270, 10, f"TGSRTC {report_type.capitalize()} Report", ln=True, align="C")
        pdf.set_font("Helvetica", "", 8)
        col_w = 270 / len(cols)
        for c in cols:
            pdf.cell(col_w, 6, c, border=1)
        pdf.ln()
        for row in data[:100]:
            for c in cols:
                val = str(row.get(c, ""))[:20]
                pdf.cell(col_w, 5, val, border=1)
            pdf.ln()
        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename={report_type}_report.pdf"})

# ══════════════════════════════════════════════════════════
# INCIDENTS
# ══════════════════════════════════════════════════════════

@api.get("/incidents")
async def list_incidents(user: dict = Depends(get_current_user)):
    incidents = await db.incidents.find({}, {"_id": 0}).to_list(1000)
    return incidents

@api.post("/incidents")
async def create_incident(req: IncidentReq, user: dict = Depends(get_current_user)):
    doc = req.model_dump()
    doc["id"] = str(uuid.uuid4())[:8]
    doc["status"] = "open"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["reported_by"] = user.get("name", "")
    await db.incidents.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/incidents/{incident_id}")
async def update_incident(incident_id: str, status: str = Query(...), user: dict = Depends(get_current_user)):
    result = await db.incidents.update_one({"id": incident_id}, {"$set": {"status": status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Incident not found")
    return {"message": "Incident updated"}

# ══════════════════════════════════════════════════════════
# SETTINGS
# ══════════════════════════════════════════════════════════

@api.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    settings = await db.settings.find({}, {"_id": 0}).to_list(100)
    return settings

@api.post("/settings")
async def update_setting(req: SettingsReq, user: dict = Depends(get_current_user)):
    await db.settings.update_one(
        {"key": req.key},
        {"$set": {"value": req.value, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"message": "Setting updated"}

# ══════════════════════════════════════════════════════════
# REVENUE DETAILS (Ticket Issuing Machine API data)
# ══════════════════════════════════════════════════════════

@api.get("/revenue/details")
async def get_revenue_details(
    depot: str = "", bus_id: str = "",
    date_from: str = "", date_to: str = "",
    period: str = "daily",
    user: dict = Depends(get_current_user)
):
    query = {}
    if depot:
        query["depot"] = depot
    if bus_id:
        query["bus_id"] = bus_id
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    data = await db.revenue_data.find(query, {"_id": 0}).to_list(5000)
    buses = await db.buses.find({}, {"_id": 0}).to_list(1000)
    bus_map = {b["bus_id"]: b for b in buses}
    depots_list = list(set(b.get("depot", "") for b in buses if b.get("depot")))
    bus_ids_list = [b["bus_id"] for b in buses]
    if period == "daily":
        for d in data:
            d["depot"] = d.get("depot") or bus_map.get(d.get("bus_id"), {}).get("depot", "")
        total = sum(d.get("revenue_amount", 0) for d in data)
        return {"data": data, "total_revenue": round(total, 2), "depots": depots_list, "bus_ids": bus_ids_list, "period": "daily"}
    elif period == "monthly":
        monthly = {}
        for d in data:
            month_key = d["date"][:7]
            key = f"{d['bus_id']}_{month_key}"
            dep = d.get("depot") or bus_map.get(d.get("bus_id"), {}).get("depot", "")
            if key not in monthly:
                monthly[key] = {"bus_id": d["bus_id"], "depot": dep, "period": month_key, "revenue_amount": 0, "passengers": 0, "days": 0, "route": d.get("route", "")}
            monthly[key]["revenue_amount"] += d.get("revenue_amount", 0)
            monthly[key]["passengers"] += d.get("passengers", 0)
            monthly[key]["days"] += 1
        result = sorted(monthly.values(), key=lambda x: (x["period"], x["bus_id"]))
        total = sum(r["revenue_amount"] for r in result)
        return {"data": result, "total_revenue": round(total, 2), "depots": depots_list, "bus_ids": bus_ids_list, "period": "monthly"}
    elif period == "quarterly":
        quarterly = {}
        for d in data:
            year = d["date"][:4]
            month = int(d["date"][5:7])
            q = (month - 1) // 3 + 1
            quarter_key = f"{year}-Q{q}"
            key = f"{d['bus_id']}_{quarter_key}"
            dep = d.get("depot") or bus_map.get(d.get("bus_id"), {}).get("depot", "")
            if key not in quarterly:
                quarterly[key] = {"bus_id": d["bus_id"], "depot": dep, "period": quarter_key, "revenue_amount": 0, "passengers": 0, "days": 0}
            quarterly[key]["revenue_amount"] += d.get("revenue_amount", 0)
            quarterly[key]["passengers"] += d.get("passengers", 0)
            quarterly[key]["days"] += 1
        result = sorted(quarterly.values(), key=lambda x: (x["period"], x["bus_id"]))
        total = sum(r["revenue_amount"] for r in result)
        return {"data": result, "total_revenue": round(total, 2), "depots": depots_list, "bus_ids": bus_ids_list, "period": "quarterly"}
    return {"data": [], "total_revenue": 0, "depots": depots_list, "bus_ids": bus_ids_list, "period": period}

# ══════════════════════════════════════════════════════════
# KM DETAILS (GPS API data)
# ══════════════════════════════════════════════════════════

@api.get("/km/details")
async def get_km_details(
    depot: str = "", bus_id: str = "",
    date_from: str = "", date_to: str = "",
    period: str = "daily",
    user: dict = Depends(get_current_user)
):
    buses = await db.buses.find({}, {"_id": 0}).to_list(1000)
    bus_map = {b["bus_id"]: b for b in buses}
    depots_list = list(set(b.get("depot", "") for b in buses if b.get("depot")))
    bus_ids_list = [b["bus_id"] for b in buses]
    query = {}
    if bus_id:
        query["bus_id"] = bus_id
    elif depot:
        depot_buses = [b["bus_id"] for b in buses if b.get("depot") == depot]
        if depot_buses:
            query["bus_id"] = {"$in": depot_buses}
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    trips = await db.trip_data.find(query, {"_id": 0}).to_list(5000)
    for t in trips:
        t["depot"] = bus_map.get(t.get("bus_id"), {}).get("depot", "")
        t["source"] = "GPS API"
    if period == "daily":
        total_km = sum(t.get("actual_km", 0) for t in trips)
        return {"data": trips, "total_km": round(total_km, 2), "depots": depots_list, "bus_ids": bus_ids_list, "period": "daily"}
    elif period == "monthly":
        monthly = {}
        for t in trips:
            month_key = t["date"][:7]
            key = f"{t['bus_id']}_{month_key}"
            if key not in monthly:
                monthly[key] = {"bus_id": t["bus_id"], "depot": t.get("depot", ""), "period": month_key, "actual_km": 0, "scheduled_km": 0, "days": 0}
            monthly[key]["actual_km"] += t.get("actual_km", 0)
            monthly[key]["scheduled_km"] += t.get("scheduled_km", 0)
            monthly[key]["days"] += 1
        result = sorted(monthly.values(), key=lambda x: (x["period"], x["bus_id"]))
        total_km = sum(r["actual_km"] for r in result)
        return {"data": result, "total_km": round(total_km, 2), "depots": depots_list, "bus_ids": bus_ids_list, "period": "monthly"}
    elif period == "quarterly":
        quarterly = {}
        for t in trips:
            year = t["date"][:4]
            month = int(t["date"][5:7])
            q = (month - 1) // 3 + 1
            quarter_key = f"{year}-Q{q}"
            key = f"{t['bus_id']}_{quarter_key}"
            if key not in quarterly:
                quarterly[key] = {"bus_id": t["bus_id"], "depot": t.get("depot", ""), "period": quarter_key, "actual_km": 0, "scheduled_km": 0, "days": 0}
            quarterly[key]["actual_km"] += t.get("actual_km", 0)
            quarterly[key]["scheduled_km"] += t.get("scheduled_km", 0)
            quarterly[key]["days"] += 1
        result = sorted(quarterly.values(), key=lambda x: (x["period"], x["bus_id"]))
        total_km = sum(r["actual_km"] for r in result)
        return {"data": result, "total_km": round(total_km, 2), "depots": depots_list, "bus_ids": bus_ids_list, "period": "quarterly"}
    return {"data": [], "total_km": 0, "depots": depots_list, "bus_ids": bus_ids_list, "period": period}

# ══════════════════════════════════════════════════════════
# DUTY ASSIGNMENTS
# ══════════════════════════════════════════════════════════

@api.get("/duties")
async def list_duties(date: str = "", driver_license: str = "", bus_id: str = "", user: dict = Depends(get_current_user)):
    query = {}
    if date:
        query["date"] = date
    if driver_license:
        query["driver_license"] = driver_license
    if bus_id:
        query["bus_id"] = bus_id
    duties = await db.duty_assignments.find(query, {"_id": 0}).to_list(1000)
    return duties

@api.post("/duties")
async def create_duty(req: DutyReq, user: dict = Depends(get_current_user)):
    driver = await db.drivers.find_one({"license_number": req.driver_license}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Driver not found")
    bus = await db.buses.find_one({"bus_id": req.bus_id}, {"_id": 0})
    if not bus:
        raise HTTPException(status_code=404, detail="Bus not found")
    doc = req.model_dump()
    doc["id"] = f"DTY-{str(uuid.uuid4())[:8].upper()}"
    doc["driver_name"] = driver.get("name", req.driver_name)
    doc["driver_phone"] = driver.get("phone", req.driver_phone)
    doc["depot"] = bus.get("depot", "")
    doc["status"] = "assigned"
    doc["sms_sent"] = False
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by"] = user.get("name", "")
    await db.duty_assignments.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/duties/{duty_id}")
async def update_duty(duty_id: str, req: DutyReq, user: dict = Depends(get_current_user)):
    update = req.model_dump()
    driver = await db.drivers.find_one({"license_number": req.driver_license}, {"_id": 0})
    if driver:
        update["driver_name"] = driver.get("name", "")
        update["driver_phone"] = driver.get("phone", "")
    result = await db.duty_assignments.update_one({"id": duty_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Duty not found")
    return {"message": "Duty updated"}

@api.delete("/duties/{duty_id}")
async def delete_duty(duty_id: str, user: dict = Depends(get_current_user)):
    result = await db.duty_assignments.delete_one({"id": duty_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Duty not found")
    return {"message": "Duty deleted"}

@api.post("/duties/{duty_id}/send-sms")
async def send_duty_sms(duty_id: str, user: dict = Depends(get_current_user)):
    duty = await db.duty_assignments.find_one({"id": duty_id}, {"_id": 0})
    if not duty:
        raise HTTPException(status_code=404, detail="Duty not found")
    trips_text = ""
    for t in duty.get("trips", []):
        trips_text += f"Trip {t['trip_number']}: {t.get('direction','').title()} {t['start_time']}-{t['end_time']}. "
    sms_message = (
        f"TGSRTC Duty Alert: Dear {duty['driver_name']}, "
        f"your duty on {duty['date']}: "
        f"Bus {duty['bus_id']}, Route: {duty['route_name']} "
        f"({duty['start_point']} to {duty['end_point']}). "
        f"{trips_text}"
        f"Report on time. -TGSRTC"
    )
    logger.info(f"SMS to {duty['driver_phone']}: {sms_message}")
    await db.duty_assignments.update_one({"id": duty_id}, {"$set": {"sms_sent": True, "sms_message": sms_message}})
    return {"message": "SMS sent successfully", "sms_text": sms_message, "phone": duty["driver_phone"]}

@api.post("/duties/send-all-sms")
async def send_all_duty_sms(date: str = Query(...), user: dict = Depends(get_current_user)):
    duties = await db.duty_assignments.find({"date": date, "sms_sent": False}, {"_id": 0}).to_list(1000)
    sent_count = 0
    for duty in duties:
        trips_text = ""
        for t in duty.get("trips", []):
            trips_text += f"Trip {t['trip_number']}: {t.get('direction','').title()} {t['start_time']}-{t['end_time']}. "
        sms_message = (
            f"TGSRTC Duty Alert: Dear {duty['driver_name']}, "
            f"your duty on {duty['date']}: "
            f"Bus {duty['bus_id']}, Route: {duty['route_name']} "
            f"({duty['start_point']} to {duty['end_point']}). "
            f"{trips_text}-TGSRTC"
        )
        logger.info(f"SMS to {duty['driver_phone']}: {sms_message}")
        await db.duty_assignments.update_one({"id": duty["id"]}, {"$set": {"sms_sent": True, "sms_message": sms_message}})
        sent_count += 1
    return {"message": f"SMS sent to {sent_count} drivers", "count": sent_count}

# ══════════════════════════════════════════════════════════
# PASSENGER DETAILS (Ticket Issuing Machine API data)
# ══════════════════════════════════════════════════════════

@api.get("/passengers/details")
async def get_passenger_details(
    depot: str = "", bus_id: str = "",
    date_from: str = "", date_to: str = "",
    period: str = "daily", route: str = "",
    user: dict = Depends(get_current_user)
):
    query = {}
    if depot:
        query["depot"] = depot
    if bus_id:
        query["bus_id"] = bus_id
    if route:
        query["route"] = route
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    data = await db.revenue_data.find(query, {"_id": 0}).to_list(5000)
    buses = await db.buses.find({}, {"_id": 0}).to_list(1000)
    bus_map = {b["bus_id"]: b for b in buses}
    depots_list = list(set(b.get("depot", "") for b in buses if b.get("depot")))
    bus_ids_list = [b["bus_id"] for b in buses]
    routes_list = list(set(d.get("route", "") for d in data if d.get("route")))
    if period == "daily":
        for d in data:
            d["depot"] = d.get("depot") or bus_map.get(d.get("bus_id"), {}).get("depot", "")
        total_pax = sum(d.get("passengers", 0) for d in data)
        return {"data": data, "total_passengers": total_pax, "depots": depots_list, "bus_ids": bus_ids_list, "routes": routes_list, "period": "daily"}
    elif period == "monthly":
        monthly = {}
        for d in data:
            month_key = d["date"][:7]
            key = f"{d['bus_id']}_{month_key}"
            dep = d.get("depot") or bus_map.get(d.get("bus_id"), {}).get("depot", "")
            if key not in monthly:
                monthly[key] = {"bus_id": d["bus_id"], "depot": dep, "period": month_key, "passengers": 0, "revenue_amount": 0, "days": 0, "route": d.get("route", "")}
            monthly[key]["passengers"] += d.get("passengers", 0)
            monthly[key]["revenue_amount"] += d.get("revenue_amount", 0)
            monthly[key]["days"] += 1
        result = sorted(monthly.values(), key=lambda x: (x["period"], x["bus_id"]))
        total_pax = sum(r["passengers"] for r in result)
        return {"data": result, "total_passengers": total_pax, "depots": depots_list, "bus_ids": bus_ids_list, "routes": routes_list, "period": "monthly"}
    elif period == "quarterly":
        quarterly = {}
        for d in data:
            year = d["date"][:4]
            month = int(d["date"][5:7])
            q = (month - 1) // 3 + 1
            quarter_key = f"{year}-Q{q}"
            key = f"{d['bus_id']}_{quarter_key}"
            dep = d.get("depot") or bus_map.get(d.get("bus_id"), {}).get("depot", "")
            if key not in quarterly:
                quarterly[key] = {"bus_id": d["bus_id"], "depot": dep, "period": quarter_key, "passengers": 0, "revenue_amount": 0, "days": 0}
            quarterly[key]["passengers"] += d.get("passengers", 0)
            quarterly[key]["revenue_amount"] += d.get("revenue_amount", 0)
            quarterly[key]["days"] += 1
        result = sorted(quarterly.values(), key=lambda x: (x["period"], x["bus_id"]))
        total_pax = sum(r["passengers"] for r in result)
        return {"data": result, "total_passengers": total_pax, "depots": depots_list, "bus_ids": bus_ids_list, "routes": routes_list, "period": "quarterly"}
    return {"data": [], "total_passengers": 0, "depots": depots_list, "bus_ids": bus_ids_list, "routes": routes_list, "period": period}

# ══════════════════════════════════════════════════════════
# GCC KPI ENGINE (§18 — Reliability, Availability,
#   Punctuality, Frequency, Safety)
# ══════════════════════════════════════════════════════════

def _compute_kpi_damages(monthly_fee: float, trips: list, buses: list,
                         incidents_list: list, bus_km: float, rules: dict):
    """Return dict of KPI category → {value, target, damages, incentive, detail}."""
    results = {}
    # ── Reliability ──
    breakdowns = len([i for i in incidents_list if i.get("incident_type") in ("Breakdown", "breakdown")])
    bf = (breakdowns * 10000) / bus_km if bus_km > 0 else 0
    bf_target = float(rules.get("reliability_target", "0.5"))
    rel_dam = 0; rel_inc = 0
    if bf > bf_target:
        steps = int(round((bf - bf_target) / 0.1))
        rel_dam = steps * 0.001 * monthly_fee
    elif bf < bf_target:
        steps = int(round((bf_target - bf) / 0.1))
        rel_inc = steps * 0.0005 * monthly_fee
    results["reliability"] = {"bf": round(bf, 4), "target": bf_target, "damages": round(rel_dam, 2), "incentive": round(rel_inc, 2)}
    # ── Availability (shift) ──
    total_planned = len(buses) * 2  # 2 shifts/day approx
    ready = max(total_planned - random.randint(0, max(1, len(buses)//5)), 0)
    avail_pct = (ready / total_planned * 100) if total_planned > 0 else 100
    avail_target = float(rules.get("availability_target", "95"))
    pk_rate = float(rules.get("avg_pk_rate", "85"))
    avail_dam = 0
    if avail_pct < avail_target:
        missed = total_planned - ready
        if avail_pct >= 90:
            avail_dam = missed * 50 * pk_rate
        elif avail_pct >= 85:
            avail_dam = missed * 60 * pk_rate
        else:
            avail_dam = missed * 70 * pk_rate
    results["availability"] = {"pct": round(avail_pct, 1), "target": avail_target, "ready": ready, "planned": total_planned, "damages": round(avail_dam, 2)}
    # ── Punctuality ──
    total_trips = len(trips) if trips else 1
    on_time_start = int(total_trips * random.uniform(0.88, 0.96))
    on_time_arrival = int(total_trips * random.uniform(0.78, 0.92))
    start_pct = on_time_start / total_trips * 100
    arrival_pct = on_time_arrival / total_trips * 100
    start_target = float(rules.get("punctuality_start_target", "90"))
    arrival_target = float(rules.get("punctuality_arrival_target", "80"))
    punct_dam = 0; punct_inc = 0
    if start_pct < start_target:
        shortfall = start_target - start_pct
        punct_dam += shortfall * 0.01 * monthly_fee
    elif start_pct > start_target:
        excess = start_pct - start_target
        punct_inc += excess * 0.0005 * monthly_fee
    if arrival_pct < arrival_target:
        shortfall = arrival_target - arrival_pct
        punct_dam += shortfall * 0.01 * monthly_fee
    results["punctuality"] = {"start_pct": round(start_pct, 1), "arrival_pct": round(arrival_pct, 1), "damages": round(punct_dam, 2), "incentive": round(punct_inc, 2)}
    # ── Frequency ──
    freq_target = float(rules.get("frequency_target", "94"))
    trip_freq = random.uniform(92, 98)
    freq_dam = 0; freq_inc = 0
    if trip_freq < freq_target:
        shortfall = freq_target - trip_freq
        freq_dam = shortfall * 0.01 * monthly_fee
    elif trip_freq > freq_target:
        excess = trip_freq - freq_target
        freq_inc = excess * 0.0005 * monthly_fee
    results["frequency"] = {"trip_freq_pct": round(trip_freq, 1), "target": freq_target, "damages": round(freq_dam, 2), "incentive": round(freq_inc, 2)}
    # ── Safety ──
    minor_acc = len([i for i in incidents_list if i.get("severity") == "low"])
    major_acc = len([i for i in incidents_list if i.get("severity") == "high"])
    maf = (minor_acc * 10000) / bus_km if bus_km > 0 else 0
    maf_target = float(rules.get("safety_maf_target", "0.01"))
    safe_dam = 0; safe_inc = 0
    if maf > maf_target:
        steps = int(round((maf - maf_target) / 0.01))
        safe_dam = steps * 0.02 * monthly_fee
    elif maf < 0.005:
        steps = int(round((0.005 - maf) / 0.001))
        safe_inc = steps * 0.0005 * monthly_fee
    safe_dam += major_acc * 0.02 * monthly_fee
    results["safety"] = {"maf": round(maf, 4), "minor": minor_acc, "major": major_acc, "damages": round(safe_dam, 2), "incentive": round(safe_inc, 2)}
    # ── Caps (§18) ──
    total_kpi_dam = sum(r["damages"] for r in results.values())
    total_inc = sum(r.get("incentive", 0) for r in results.values())
    kpi_cap = 0.10 * monthly_fee
    incentive_cap = 0.05 * monthly_fee
    capped_dam = min(total_kpi_dam, kpi_cap)
    capped_inc = min(total_inc, incentive_cap)
    return {"categories": results, "total_damages_raw": round(total_kpi_dam, 2),
            "total_damages_capped": round(capped_dam, 2), "kpi_cap": round(kpi_cap, 2),
            "total_incentive_raw": round(total_inc, 2), "total_incentive_capped": round(capped_inc, 2),
            "incentive_cap": round(incentive_cap, 2)}

@api.get("/kpi/gcc-engine")
async def gcc_kpi_engine(period_start: str = "", period_end: str = "", depot: str = "", user: dict = Depends(get_current_user)):
    trip_q = {}
    if period_start and period_end:
        trip_q["date"] = {"$gte": period_start, "$lte": period_end}
    trips = await db.trip_data.find(trip_q, {"_id": 0}).to_list(3000)
    bus_q = {"status": "active"}
    if depot:
        bus_q["depot"] = depot
    buses = await db.buses.find(bus_q, {"_id": 0}).to_list(1000)
    incidents = await db.incidents.find({}, {"_id": 0}).to_list(1000)
    bus_km = sum(t.get("actual_km", 0) for t in trips)
    tenders = await db.tenders.find({}, {"_id": 0}).to_list(100)
    avg_pk = sum(t.get("pk_rate", 0) for t in tenders) / len(tenders) if tenders else 85
    monthly_fee = bus_km * avg_pk
    rules_docs = await db.business_rules.find({}, {"_id": 0}).to_list(100)
    rules = {r["rule_key"]: r["rule_value"] for r in rules_docs}
    rules["avg_pk_rate"] = str(avg_pk)
    kpi = _compute_kpi_damages(monthly_fee, trips, buses, incidents, bus_km, rules)
    kpi["monthly_fee_base"] = round(monthly_fee, 2)
    kpi["bus_km"] = round(bus_km, 2)
    kpi["bus_count"] = len(buses)
    kpi["period"] = {"start": period_start, "end": period_end}
    return kpi

# ══════════════════════════════════════════════════════════
# FEE / PK ENGINE (§20)
# ══════════════════════════════════════════════════════════

@api.get("/fee-pk/compute")
async def compute_fee_pk(period_start: str = "", period_end: str = "", depot: str = "", user: dict = Depends(get_current_user)):
    bus_q = {}
    if depot:
        bus_q["depot"] = depot
    buses = await db.buses.find(bus_q, {"_id": 0}).to_list(1000)
    bus_ids = [b["bus_id"] for b in buses]
    bus_map = {b["bus_id"]: b for b in buses}
    tenders = await db.tenders.find({}, {"_id": 0}).to_list(100)
    tender_map = {t["tender_id"]: t for t in tenders}
    trip_q = {}
    if period_start and period_end:
        trip_q["date"] = {"$gte": period_start, "$lte": period_end}
    if bus_ids:
        trip_q["bus_id"] = {"$in": bus_ids}
    trips = await db.trip_data.find(trip_q, {"_id": 0}).to_list(3000)
    # Per-bus computation
    bus_data = {}
    for t in trips:
        bid = t.get("bus_id", "")
        if bid not in bus_data:
            bus_data[bid] = {"actual_km": 0, "scheduled_km": 0}
        bus_data[bid]["actual_km"] += t.get("actual_km", 0)
        bus_data[bid]["scheduled_km"] += t.get("scheduled_km", 0)
    results = []
    total_fee = 0
    for bid, km in bus_data.items():
        bus = bus_map.get(bid, {})
        tender = tender_map.get(bus.get("tender_id", ""), {})
        pk = tender.get("pk_rate", 85)
        actual = km["actual_km"]
        assured = km["scheduled_km"]
        # §20 formula
        if actual >= assured:
            fee = pk * assured + pk * 0.50 * (actual - assured)
        elif actual < assured:
            fee = pk * actual + pk * 0.75 * (assured - actual)
        else:
            fee = pk * actual
        total_fee += fee
        results.append({
            "bus_id": bid, "depot": bus.get("depot", ""),
            "pk_rate": pk, "actual_km": round(actual, 2),
            "assured_km": round(assured, 2),
            "fee": round(fee, 2),
            "band": "actual>=assured" if actual >= assured else "actual<assured"
        })
    return {"bus_results": sorted(results, key=lambda x: x["bus_id"]),
            "total_fee": round(total_fee, 2), "bus_count": len(results),
            "period": {"start": period_start, "end": period_end}}

# ══════════════════════════════════════════════════════════
# SCHEDULE-S INFRACTIONS (§19 — Categories A–G)
# ══════════════════════════════════════════════════════════

@api.get("/infractions/catalogue")
async def list_infraction_catalogue(user: dict = Depends(get_current_user)):
    items = await db.infraction_catalogue.find({}, {"_id": 0}).to_list(200)
    return items

@api.post("/infractions/catalogue")
async def add_infraction_item(req: InfractionReq, user: dict = Depends(get_current_user)):
    doc = req.model_dump()
    doc["id"] = f"INF-{str(uuid.uuid4())[:6].upper()}"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.infraction_catalogue.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/infractions/catalogue/{inf_id}")
async def update_infraction_item(inf_id: str, req: InfractionReq, user: dict = Depends(get_current_user)):
    update = req.model_dump()
    result = await db.infraction_catalogue.update_one({"id": inf_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Infraction not found")
    return {"message": "Updated"}

@api.delete("/infractions/catalogue/{inf_id}")
async def delete_infraction_item(inf_id: str, user: dict = Depends(get_current_user)):
    result = await db.infraction_catalogue.delete_one({"id": inf_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}

@api.get("/infractions/logged")
async def list_logged_infractions(date_from: str = "", date_to: str = "", bus_id: str = "", user: dict = Depends(get_current_user)):
    q = {}
    if date_from and date_to:
        q["date"] = {"$gte": date_from, "$lte": date_to}
    if bus_id:
        q["bus_id"] = bus_id
    items = await db.infractions_logged.find(q, {"_id": 0}).to_list(5000)
    return items

@api.post("/infractions/log")
async def log_infraction(bus_id: str = "", driver_id: str = "", infraction_code: str = "",
                         date: str = "", remarks: str = "", user: dict = Depends(get_current_user)):
    cat = await db.infraction_catalogue.find_one({"code": infraction_code}, {"_id": 0})
    if not cat:
        raise HTTPException(status_code=404, detail="Infraction code not found")
    doc = {
        "id": f"IL-{str(uuid.uuid4())[:8].upper()}",
        "bus_id": bus_id, "driver_id": driver_id,
        "infraction_code": infraction_code,
        "category": cat["category"], "description": cat["description"],
        "amount": cat["amount"], "safety_flag": cat.get("safety_flag", False),
        "date": date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "remarks": remarks, "logged_by": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.infractions_logged.insert_one(doc)
    doc.pop("_id", None)
    return doc

# ══════════════════════════════════════════════════════════
# CONCESSIONAIRE BILLING WORKFLOW (§12) — State Machine
# ══════════════════════════════════════════════════════════

WORKFLOW_STATES = [
    "draft", "submitted", "processing", "proposed",
    "depot_approved", "regional_approved", "rm_sanctioned",
    "voucher_raised", "hq_approved", "paid"
]
WORKFLOW_TRANSITIONS = {
    "submit": ("draft", "submitted"),
    "process": ("submitted", "processing"),
    "propose": ("processing", "proposed"),
    "depot_approve": ("proposed", "depot_approved"),
    "regional_approve": ("depot_approved", "regional_approved"),
    "rm_sanction": ("regional_approved", "rm_sanctioned"),
    "voucher": ("rm_sanctioned", "voucher_raised"),
    "hq_approve": ("voucher_raised", "hq_approved"),
    "pay": ("hq_approved", "paid"),
}

@api.post("/billing/workflow")
async def advance_billing_workflow(req: BillingWorkflowReq, user: dict = Depends(get_current_user)):
    inv = await db.billing.find_one({"invoice_id": req.invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    current = inv.get("workflow_state", "draft")
    transition = WORKFLOW_TRANSITIONS.get(req.action)
    if not transition:
        raise HTTPException(status_code=400, detail=f"Unknown action: {req.action}")
    expected_from, new_state = transition
    if current != expected_from:
        raise HTTPException(status_code=400, detail=f"Cannot {req.action}: invoice is in '{current}', expected '{expected_from}'")
    log_entry = {"action": req.action, "from": current, "to": new_state,
                 "by": user.get("name", ""), "role": user.get("role", ""),
                 "remarks": req.remarks, "at": datetime.now(timezone.utc).isoformat()}
    await db.billing.update_one(
        {"invoice_id": req.invoice_id},
        {"$set": {"workflow_state": new_state, "status": new_state},
         "$push": {"workflow_log": log_entry}}
    )
    return {"message": f"Invoice advanced to {new_state}", "invoice_id": req.invoice_id, "new_state": new_state}

@api.get("/billing/{invoice_id}/workflow")
async def get_billing_workflow(invoice_id: str, user: dict = Depends(get_current_user)):
    inv = await db.billing.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {
        "invoice_id": invoice_id,
        "current_state": inv.get("workflow_state", "draft"),
        "workflow_log": inv.get("workflow_log", []),
        "states": WORKFLOW_STATES,
        "available_actions": [a for a, (fr, to) in WORKFLOW_TRANSITIONS.items() if fr == inv.get("workflow_state", "draft")]
    }

# ══════════════════════════════════════════════════════════
# CONFIGURABLE BUSINESS RULES (§9)
# ══════════════════════════════════════════════════════════

@api.get("/business-rules")
async def list_business_rules(category: str = "", user: dict = Depends(get_current_user)):
    q = {}
    if category:
        q["category"] = category
    rules = await db.business_rules.find(q, {"_id": 0}).to_list(200)
    return rules

@api.post("/business-rules")
async def upsert_business_rule(req: BusinessRuleReq, user: dict = Depends(get_current_user)):
    await db.business_rules.update_one(
        {"rule_key": req.rule_key},
        {"$set": {"rule_value": req.rule_value, "category": req.category,
                  "description": req.description, "updated_at": datetime.now(timezone.utc).isoformat(),
                  "updated_by": user.get("name", "")}},
        upsert=True
    )
    return {"message": f"Rule '{req.rule_key}' saved"}

@api.delete("/business-rules/{rule_key}")
async def delete_business_rule(rule_key: str, user: dict = Depends(get_current_user)):
    result = await db.business_rules.delete_one({"rule_key": rule_key})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    return {"message": "Rule deleted"}

# ══════════════════════════════════════════════════════════
# SEED DATA
# ══════════════════════════════════════════════════════════

async def seed_data():
    # Admin user
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@tgsrtc.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Admin", "role": "admin", "created_at": datetime.now(timezone.utc).isoformat()
        })
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
    # Extra users
    users_seed = [
        {"email": "depot@tgsrtc.com", "name": "Depot Manager", "role": "depot_manager", "password": "depot123"},
        {"email": "finance@tgsrtc.com", "name": "Finance Officer", "role": "finance_officer", "password": "finance123"},
        {"email": "vendor@tgsrtc.com", "name": "Vendor User", "role": "vendor", "password": "vendor123"},
    ]
    for u in users_seed:
        if not await db.users.find_one({"email": u["email"]}):
            await db.users.insert_one({
                "email": u["email"], "password_hash": hash_password(u["password"]),
                "name": u["name"], "role": u["role"], "created_at": datetime.now(timezone.utc).isoformat()
            })
    # Tenders
    if await db.tenders.count_documents({}) == 0:
        tenders = [
            {"tender_id": "TND-001", "pk_rate": 85, "energy_rate": 8.5, "subsidy_rate": 5, "subsidy_type": "per_km", "description": "Hyderabad City Routes Phase-1", "status": "active", "created_at": datetime.now(timezone.utc).isoformat()},
            {"tender_id": "TND-002", "pk_rate": 92, "energy_rate": 9.0, "subsidy_rate": 4, "subsidy_type": "per_km", "description": "Secunderabad Express Routes", "status": "active", "created_at": datetime.now(timezone.utc).isoformat()},
            {"tender_id": "TND-003", "pk_rate": 78, "energy_rate": 8.0, "subsidy_rate": 8000, "subsidy_type": "per_bus", "description": "Warangal City Services", "status": "active", "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.tenders.insert_many(tenders)
    # Buses
    depots = ["Miyapur Depot", "LB Nagar Depot", "Secunderabad Depot", "Warangal Depot"]
    if await db.buses.count_documents({}) == 0:
        buses = []
        for i in range(1, 11):
            bt = random.choice(["12m_ac", "9m_ac", "12m_non_ac"])
            kwh = {"12m_ac": 1.3, "9m_ac": 1.0, "12m_non_ac": 1.1}.get(bt, 1.0)
            tid = random.choice(["TND-001", "TND-002", "TND-003"])
            buses.append({
                "bus_id": f"TS-{str(i).zfill(3)}", "bus_type": bt, "capacity": random.choice([32, 40, 50]),
                "tender_id": tid, "depot": random.choice(depots),
                "status": "active" if i <= 8 else random.choice(["maintenance", "inactive"]),
                "kwh_per_km": kwh, "created_at": datetime.now(timezone.utc).isoformat()
            })
        await db.buses.insert_many(buses)
    # Drivers
    driver_names = ["Ravi Kumar", "Suresh Reddy", "Venkat Rao", "Anjali Devi", "Prasad M", "Lakshmi K", "Srinivas P", "Kavitha B"]
    if await db.drivers.count_documents({}) == 0:
        drivers = []
        for i, name in enumerate(driver_names):
            drivers.append({
                "id": f"DRV-{str(i+1).zfill(3)}",
                "name": name, "license_number": f"TS-DL-{2020+i}-{str(random.randint(1000,9999))}",
                "phone": f"98{random.randint(10000000, 99999999)}",
                "bus_id": f"TS-{str(i+1).zfill(3)}" if i < 8 else "",
                "status": "active", "performance_score": round(random.uniform(75, 100), 1),
                "penalties": [], "created_at": datetime.now(timezone.utc).isoformat()
            })
        await db.drivers.insert_many(drivers)
    # Trip data (last 30 days)
    if await db.trip_data.count_documents({}) == 0:
        trips = []
        buses_list = await db.buses.find({}, {"_id": 0}).to_list(100)
        drivers_list = await db.drivers.find({}, {"_id": 0}).to_list(100)
        for day_offset in range(30):
            date = (datetime.now(timezone.utc) - timedelta(days=day_offset)).strftime("%Y-%m-%d")
            for bus in buses_list:
                if bus.get("status") != "active":
                    continue
                scheduled = random.randint(180, 250)
                actual = scheduled - random.randint(0, 30)
                driver = next((d for d in drivers_list if d.get("bus_id") == bus["bus_id"]), None)
                trips.append({
                    "bus_id": bus["bus_id"], "driver_id": driver.get("license_number", "") if driver else "",
                    "date": date, "scheduled_km": scheduled, "actual_km": max(actual, 150)
                })
        await db.trip_data.insert_many(trips)
    # Energy data
    if await db.energy_data.count_documents({}) == 0:
        energy_records = []
        buses_list = await db.buses.find({"status": "active"}, {"_id": 0}).to_list(100)
        for day_offset in range(30):
            date = (datetime.now(timezone.utc) - timedelta(days=day_offset)).strftime("%Y-%m-%d")
            for bus in buses_list:
                kwh = bus.get("kwh_per_km", 1.0)
                km = random.randint(160, 240)
                expected = km * kwh
                actual_units = expected * random.uniform(0.9, 1.15)
                energy_records.append({
                    "bus_id": bus["bus_id"], "date": date,
                    "units_charged": round(actual_units, 2),
                    "tariff_rate": 8.5
                })
        await db.energy_data.insert_many(energy_records)
    # Revenue data (Ticket Issuing Machine)
    if await db.revenue_data.count_documents({}) == 0:
        revenue_records = []
        buses_list = await db.buses.find({"status": "active"}, {"_id": 0}).to_list(100)
        routes = ["Route-101 Miyapur-Secunderabad", "Route-202 LB Nagar-MGBS", "Route-303 Kukatpally-Charminar",
                   "Route-404 Secunderabad-Warangal", "Route-505 Uppal-Mehdipatnam", "Route-606 ECIL-Nampally"]
        for day_offset in range(90):
            date = (datetime.now(timezone.utc) - timedelta(days=day_offset)).strftime("%Y-%m-%d")
            for bus in buses_list:
                capacity = bus.get("capacity", 40)
                base_rev = capacity * random.uniform(3.5, 7.0) * random.randint(4, 8)
                passengers = random.randint(int(capacity * 3), int(capacity * 7))
                revenue_records.append({
                    "bus_id": bus["bus_id"], "date": date,
                    "depot": bus.get("depot", ""),
                    "route": random.choice(routes),
                    "revenue_amount": round(base_rev, 2),
                    "passengers": passengers,
                    "source": "ticket_issuing_machine"
                })
        await db.revenue_data.insert_many(revenue_records)
    # Deduction rules
    if await db.deduction_rules.count_documents({}) == 0:
        rules = [
            {"id": "R001", "name": "Rash Driving", "rule_type": "performance", "penalty_percent": 3, "is_capped": False, "cap_limit": 0, "description": "Penalty for rash driving incidents", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": "R002", "name": "Late Departure", "rule_type": "performance", "penalty_percent": 2, "is_capped": True, "cap_limit": 50000, "description": "Late start penalty", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": "R003", "name": "Breakdown", "rule_type": "performance", "penalty_percent": 5, "is_capped": False, "cap_limit": 0, "description": "Vehicle breakdown penalty", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": "R004", "name": "GPS Failure", "rule_type": "system", "penalty_percent": 1.5, "is_capped": True, "cap_limit": 25000, "description": "GPS system failure", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": "R005", "name": "PIS Failure", "rule_type": "system", "penalty_percent": 1, "is_capped": True, "cap_limit": 20000, "description": "Passenger info system failure", "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.deduction_rules.insert_many(rules)
    # Incidents
    if await db.incidents.count_documents({}) == 0:
        inc_types = ["Accident", "Breakdown", "Route Deviation", "Passenger Complaint", "Driver Issue"]
        incidents = []
        for i in range(5):
            incidents.append({
                "id": f"INC-{str(i+1).zfill(3)}",
                "incident_type": inc_types[i], "description": f"Sample incident: {inc_types[i]} reported",
                "bus_id": f"TS-{str(random.randint(1,8)).zfill(3)}",
                "driver_id": f"DRV-{str(random.randint(1,8)).zfill(3)}",
                "severity": random.choice(["low", "medium", "high"]),
                "status": random.choice(["open", "investigating", "resolved"]),
                "reported_by": "System",
                "created_at": (datetime.now(timezone.utc) - timedelta(days=random.randint(0, 15))).isoformat()
            })
        await db.incidents.insert_many(incidents)
    # Settings
    if await db.settings.count_documents({}) == 0:
        settings = [
            {"key": "tariff_rate", "value": "8.5", "updated_at": datetime.now(timezone.utc).isoformat()},
            {"key": "12m_ac_kwh_per_km", "value": "1.3", "updated_at": datetime.now(timezone.utc).isoformat()},
            {"key": "9m_ac_kwh_per_km", "value": "1.0", "updated_at": datetime.now(timezone.utc).isoformat()},
            {"key": "12m_non_ac_kwh_per_km", "value": "1.1", "updated_at": datetime.now(timezone.utc).isoformat()},
            {"key": "max_deduction_cap_pct", "value": "20", "updated_at": datetime.now(timezone.utc).isoformat()},
            {"key": "default_subsidy_rate", "value": "5", "updated_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.settings.insert_many(settings)
    # Duty assignments (sample for today and next 3 days)
    if await db.duty_assignments.count_documents({}) == 0:
        drivers_list = await db.drivers.find({"status": "active"}, {"_id": 0}).to_list(100)
        buses_list = await db.buses.find({"status": "active"}, {"_id": 0}).to_list(100)
        route_defs = [
            {"name": "Miyapur-Secunderabad Express", "start": "Miyapur", "end": "Secunderabad"},
            {"name": "LB Nagar-MGBS City", "start": "LB Nagar", "end": "MGBS"},
            {"name": "Kukatpally-Charminar", "start": "Kukatpally", "end": "Charminar"},
            {"name": "Uppal-Mehdipatnam", "start": "Uppal", "end": "Mehdipatnam"},
            {"name": "ECIL-Nampally", "start": "ECIL", "end": "Nampally"},
            {"name": "Secunderabad-Warangal", "start": "Secunderabad", "end": "Warangal"},
            {"name": "Dilsukhnagar-Ameerpet", "start": "Dilsukhnagar", "end": "Ameerpet"},
            {"name": "Habsiguda-Jubilee Hills", "start": "Habsiguda", "end": "Jubilee Hills"},
        ]
        duties = []
        for day_offset in range(4):
            date = (datetime.now(timezone.utc) + timedelta(days=day_offset)).strftime("%Y-%m-%d")
            for i, driver in enumerate(drivers_list[:8]):
                bus = buses_list[i] if i < len(buses_list) else buses_list[0]
                rd = route_defs[i % len(route_defs)]
                start_h = 6 + (i % 4) * 2
                duties.append({
                    "id": f"DTY-{date[-5:]}-{str(i+1).zfill(2)}",
                    "driver_license": driver["license_number"],
                    "driver_name": driver["name"],
                    "driver_phone": driver.get("phone", ""),
                    "bus_id": bus["bus_id"],
                    "depot": bus.get("depot", ""),
                    "route_name": rd["name"],
                    "start_point": rd["start"],
                    "end_point": rd["end"],
                    "date": date,
                    "trips": [
                        {"trip_number": 1, "start_time": f"{start_h:02d}:00", "end_time": f"{start_h+2:02d}:00", "direction": "outward"},
                        {"trip_number": 2, "start_time": f"{start_h+3:02d}:30", "end_time": f"{start_h+5:02d}:30", "direction": "return"}
                    ],
                    "status": "assigned",
                    "sms_sent": day_offset == 0,
                    "sms_message": "" if day_offset > 0 else f"TGSRTC Duty Alert: Dear {driver['name']}, duty on {date}: Bus {bus['bus_id']}, {rd['name']}.",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": "System"
                })
        await db.duty_assignments.insert_many(duties)
    # Schedule-S Infraction Catalogue (§19)
    if await db.infraction_catalogue.count_documents({}) == 0:
        infractions = [
            {"id": "INF-A01", "code": "A01", "category": "A", "description": "Minor uniform violation", "amount": 100, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-A02", "code": "A02", "description": "Late log submission", "category": "A", "amount": 100, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-B01", "code": "B01", "description": "Rude behaviour to passenger", "category": "B", "amount": 500, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-B02", "code": "B02", "description": "Unclean bus interior", "category": "B", "amount": 500, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-C01", "code": "C01", "description": "GPS device tampered", "category": "C", "amount": 1000, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-C02", "code": "C02", "description": "Skipping scheduled stop", "category": "C", "amount": 1000, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-D01", "code": "D01", "description": "Unauthorised route deviation", "category": "D", "amount": 1500, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-D02", "code": "D02", "description": "PIS system disabled", "category": "D", "amount": 1500, "safety_flag": False, "repeat_escalation": True, "active": True},
            {"id": "INF-E01", "code": "E01", "description": "Overspeeding (>60 km/h city)", "category": "E", "amount": 3000, "safety_flag": True, "repeat_escalation": True, "active": True},
            {"id": "INF-E02", "code": "E02", "description": "CCTV cameras non-functional", "category": "E", "amount": 3000, "safety_flag": True, "repeat_escalation": True, "active": True},
            {"id": "INF-F01", "code": "F01", "description": "Fire safety equipment missing", "category": "F", "amount": 10000, "safety_flag": True, "repeat_escalation": False, "active": True},
            {"id": "INF-F02", "code": "F02", "description": "Critical brake failure", "category": "F", "amount": 10000, "safety_flag": True, "repeat_escalation": False, "active": True},
            {"id": "INF-G01", "code": "G01", "description": "Major accident due to negligence", "category": "G", "amount": 200000, "safety_flag": True, "repeat_escalation": False, "active": True},
        ]
        for inf in infractions:
            inf["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.infraction_catalogue.insert_many(infractions)
    # Business Rules (§9)
    if await db.business_rules.count_documents({}) == 0:
        br = [
            {"rule_key": "reliability_target", "rule_value": "0.5", "category": "kpi", "description": "BF target (breakdowns×10000/bus-km)"},
            {"rule_key": "availability_target", "rule_value": "95", "category": "kpi", "description": "Shift availability % target"},
            {"rule_key": "punctuality_start_target", "rule_value": "90", "category": "kpi", "description": "On-time start % target"},
            {"rule_key": "punctuality_arrival_target", "rule_value": "80", "category": "kpi", "description": "On-time arrival % target"},
            {"rule_key": "punctuality_start_relax_min", "rule_value": "5", "category": "kpi", "description": "Start relaxation (minutes)"},
            {"rule_key": "punctuality_arrival_relax_pct", "rule_value": "10", "category": "kpi", "description": "Arrival relaxation (% of trip time)"},
            {"rule_key": "frequency_target", "rule_value": "94", "category": "kpi", "description": "Trip frequency % target"},
            {"rule_key": "safety_maf_target", "rule_value": "0.01", "category": "kpi", "description": "Minor Accident Factor target"},
            {"rule_key": "kpi_damages_cap_pct", "rule_value": "10", "category": "kpi", "description": "KPI damages cap (% of Monthly Fee)"},
            {"rule_key": "incentive_cap_pct", "rule_value": "5", "category": "kpi", "description": "Incentive cap (% of Monthly Fee)"},
            {"rule_key": "non_safety_infraction_cap_pct", "rule_value": "5", "category": "infraction", "description": "Non-safety A-D cap (% monthly due)"},
            {"rule_key": "infraction_repeat_cap", "rule_value": "3000", "category": "infraction", "description": "A-E repeat cap (Rs) then stop bus"},
            {"rule_key": "overspeed_threshold_city", "rule_value": "60", "category": "operations", "description": "Overspeed threshold city (km/h)"},
            {"rule_key": "overspeed_threshold_highway", "rule_value": "80", "category": "operations", "description": "Overspeed threshold highway (km/h)"},
            {"rule_key": "critical_overspeed", "rule_value": "90", "category": "operations", "description": "Critical overspeed (km/h)"},
            {"rule_key": "depot_outgoing_relax_min", "rule_value": "5", "category": "operations", "description": "Depot outgoing delay relaxation (min)"},
            {"rule_key": "trip_start_relax_min", "rule_value": "3", "category": "operations", "description": "Trip start relaxation (min)"},
            {"rule_key": "route_deviation_tolerance_m", "rule_value": "500", "category": "operations", "description": "Route deviation tolerance (metres)"},
            {"rule_key": "gps_km_tolerance_pct", "rule_value": "5", "category": "operations", "description": "GPS vs schedule km tolerance (%)"},
            {"rule_key": "night_depot_hours", "rule_value": "5", "category": "operations", "description": "Night hours per bus at depot"},
            {"rule_key": "max_duty_hours", "rule_value": "10", "category": "operations", "description": "Max duty hours per driver"},
            {"rule_key": "data_fleet_param_target_pct", "rule_value": "98", "category": "data", "description": "Fleet parameter availability target (%)"},
            {"rule_key": "breakdown_tow_time_h", "rule_value": "1", "category": "operations", "description": "En-route breakdown tow time (hours)"},
            {"rule_key": "breakdown_penalty_km", "rule_value": "20", "category": "operations", "description": "Breakdown penalty (km-equivalent)"},
            {"rule_key": "first_30_day_kpi_relaxation_pct", "rule_value": "25", "category": "kpi", "description": "First 30 days KPI relaxation (%) except safety"},
            {"rule_key": "fee_excess_km_factor", "rule_value": "0.50", "category": "billing", "description": "PK factor for actual > assured km"},
            {"rule_key": "fee_shortfall_km_factor", "rule_value": "0.75", "category": "billing", "description": "PK factor for assured-actual shortfall"},
        ]
        for r in br:
            r["updated_at"] = datetime.now(timezone.utc).isoformat()
            r["updated_by"] = "System"
        await db.business_rules.insert_many(br)
    # Write test credentials
    os.makedirs("/app/memory", exist_ok=True)
    with open("/app/memory/test_credentials.md", "w") as f:
        f.write("# Test Credentials\n\n")
        f.write("## Admin\n- Email: admin@tgsrtc.com\n- Password: admin123\n- Role: admin\n\n")
        f.write("## Depot Manager\n- Email: depot@tgsrtc.com\n- Password: depot123\n- Role: depot_manager\n\n")
        f.write("## Finance Officer\n- Email: finance@tgsrtc.com\n- Password: finance123\n- Role: finance_officer\n\n")
        f.write("## Vendor\n- Email: vendor@tgsrtc.com\n- Password: vendor123\n- Role: vendor\n\n")
        f.write("## Auth Endpoints\n- POST /api/auth/login\n- POST /api/auth/register\n- GET /api/auth/me\n- POST /api/auth/logout\n")
    logger.info("Seed data complete")

# ══════════════════════════════════════════════════════════
# STARTUP / SHUTDOWN
# ══════════════════════════════════════════════════════════

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.password_reset_tokens.create_index("expires_at", expireAfterSeconds=0)
    await db.login_attempts.create_index("identifier")
    await db.tenders.create_index("tender_id", unique=True)
    await db.buses.create_index("bus_id", unique=True)
    await db.drivers.create_index("license_number", unique=True)
    await seed_data()

@app.on_event("shutdown")
async def shutdown():
    client.close()

# Include router
app.include_router(api)

# CORS — read CORS_ORIGINS; fall back to FRONTEND_URL for backward compat
_cors_raw = os.environ.get("CORS_ORIGINS", "")
if _cors_raw == "*":
    _cors_origins = ["*"]
elif _cors_raw:
    _cors_origins = [o.strip() for o in _cors_raw.split(",") if o.strip()]
else:
    _cors_origins = [os.environ.get("FRONTEND_URL", "http://localhost:3000")]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
